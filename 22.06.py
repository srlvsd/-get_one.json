import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_interval


workbook = openpyxl.load_workbook('data.xlsx')
worksheet = workbook.active

print(workbook)


value = worksheet.cell(2, 2).value
print(value)


row1_arr=[]
row2_arr=[]
row3_arr=[]

for i in range (1, worksheet.max_column+1):
    row1_arr.append(worksheet.cell(2, i).value)
    row2_arr.append(worksheet.cell(4, i).value)
    row3_arr.append(worksheet.cell(6, i).value)

print(row1_arr, row2_arr,row3_arr)
